from fastapi import APIRouter, HTTPException, Request, Query
from fastapi.responses import StreamingResponse
from auth import require_admin, db
from datetime import datetime, timezone, timedelta
from typing import Optional, List
from pydantic import BaseModel
import io
import csv

router = APIRouter(prefix="/api/admin/customers", tags=["customers"])


class CustomerStatusUpdate(BaseModel):
    is_active: bool


@router.get("")
async def get_customers(
    request: Request,
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    search: Optional[str] = Query(None, description="Search by name, email, or phone"),
    from_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    to_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    status: Optional[str] = Query(None, description="Filter by status: active, inactive, all")
):
    """
    Get paginated list of customers with search and filters.
    Uses aggregation pipeline for performance.
    Admin only.
    """
    await require_admin(request)
    
    # Build match filter - exclude admins and soft-deleted users
    match_filter = {
        "role": {"$ne": "admin"},
        "$or": [
            {"is_deleted": {"$exists": False}},
            {"is_deleted": False}
        ]
    }
    
    # Search filter
    if search:
        search_regex = {"$regex": search, "$options": "i"}
        match_filter["$and"] = match_filter.get("$and", []) + [{
            "$or": [
                {"name": search_regex},
                {"email": search_regex},
                {"phone": search_regex}
            ]
        }]
    
    # Date filter on user creation
    if from_date:
        start = datetime.strptime(from_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
        match_filter["created_at"] = {"$gte": start}
    
    if to_date:
        end = datetime.strptime(to_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59, tzinfo=timezone.utc)
        if "created_at" in match_filter:
            match_filter["created_at"]["$lte"] = end
        else:
            match_filter["created_at"] = {"$lte": end}
    
    # Status filter
    if status == "active":
        match_filter["$and"] = match_filter.get("$and", []) + [{
            "$or": [
                {"is_active": {"$exists": False}},
                {"is_active": True}
            ]
        }]
    elif status == "inactive":
        match_filter["is_active"] = False
    
    # Aggregation pipeline to join with orders for stats
    pipeline = [
        {"$match": match_filter},
        # Lookup orders for each customer
        {"$lookup": {
            "from": "orders",
            "localField": "user_id",
            "foreignField": "user_id",
            "as": "orders"
        }},
        # Calculate customer stats
        {"$addFields": {
            "total_orders": {"$size": "$orders"},
            "total_spend": {
                "$sum": {
                    "$map": {
                        "input": {
                            "$filter": {
                                "input": "$orders",
                                "as": "order",
                                "cond": {"$ne": ["$$order.order_status", "cancelled"]}
                            }
                        },
                        "as": "o",
                        "in": "$$o.total"
                    }
                }
            },
            "last_order_date": {
                "$max": "$orders.created_at"
            },
            "is_active": {"$ifNull": ["$is_active", True]}
        }},
        # Project only needed fields
        {"$project": {
            "_id": 0,
            "user_id": 1,
            "name": 1,
            "email": 1,
            "phone": 1,
            "auth_provider": 1,
            "created_at": 1,
            "is_active": 1,
            "total_orders": 1,
            "total_spend": 1,
            "last_order_date": 1
        }},
        # Sort by most recent first
        {"$sort": {"created_at": -1}},
        # Pagination with facet
        {"$facet": {
            "customers": [
                {"$skip": (page - 1) * per_page},
                {"$limit": per_page}
            ],
            "total_count": [
                {"$count": "count"}
            ]
        }}
    ]
    
    result = await db.users.aggregate(pipeline).to_list(1)
    
    if not result:
        return {
            "customers": [],
            "total": 0,
            "page": page,
            "per_page": per_page,
            "total_pages": 0
        }
    
    data = result[0]
    customers = data.get("customers", [])
    total = data["total_count"][0]["count"] if data["total_count"] else 0
    total_pages = (total + per_page - 1) // per_page
    
    return {
        "customers": customers,
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": total_pages
    }


@router.get("/{customer_id}")
async def get_customer_detail(customer_id: str, request: Request):
    """
    Get detailed customer information including profile, addresses, orders, returns, and financial summary.
    Admin only.
    """
    await require_admin(request)
    
    # Get customer base info
    customer = await db.users.find_one(
        {"user_id": customer_id, "role": {"$ne": "admin"}},
        {"_id": 0, "password": 0}
    )
    
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    # Get orders with aggregation for financial summary
    orders_pipeline = [
        {"$match": {"user_id": customer_id}},
        {"$sort": {"created_at": -1}},
        {"$project": {
            "_id": 0,
            "order_id": 1,
            "order_status": 1,
            "payment_status": 1,
            "payment_method": 1,
            "total": 1,
            "items": 1,
            "created_at": 1,
            "tracking_number": 1,
            "return_status": 1
        }}
    ]
    
    orders = await db.orders.aggregate(orders_pipeline).to_list(100)
    
    # Get returns
    returns = await db.returns.find(
        {"user_id": customer_id},
        {"_id": 0}
    ).sort("created_at", -1).to_list(50)
    
    # Calculate financial summary
    financial_pipeline = [
        {"$match": {"user_id": customer_id, "order_status": {"$ne": "cancelled"}}},
        {"$group": {
            "_id": None,
            "total_orders": {"$sum": 1},
            "total_spend": {"$sum": "$total"},
            "total_items": {"$sum": {"$size": "$items"}},
            "avg_order_value": {"$avg": "$total"},
            "delivered_orders": {
                "$sum": {"$cond": [{"$eq": ["$order_status", "delivered"]}, 1, 0]}
            },
            "cod_orders": {
                "$sum": {"$cond": [{"$eq": ["$payment_method", "cod"]}, 1, 0]}
            },
            "online_orders": {
                "$sum": {"$cond": [{"$eq": ["$payment_method", "razorpay"]}, 1, 0]}
            }
        }}
    ]
    
    financial_result = await db.orders.aggregate(financial_pipeline).to_list(1)
    
    financial_summary = {
        "total_orders": 0,
        "total_spend": 0,
        "total_items": 0,
        "avg_order_value": 0,
        "delivered_orders": 0,
        "cod_orders": 0,
        "online_orders": 0
    }
    
    if financial_result:
        fin = financial_result[0]
        financial_summary = {
            "total_orders": fin.get("total_orders", 0),
            "total_spend": round(fin.get("total_spend", 0), 2),
            "total_items": fin.get("total_items", 0),
            "avg_order_value": round(fin.get("avg_order_value", 0), 2),
            "delivered_orders": fin.get("delivered_orders", 0),
            "cod_orders": fin.get("cod_orders", 0),
            "online_orders": fin.get("online_orders", 0)
        }
    
    # Count returns
    return_counts = await db.returns.aggregate([
        {"$match": {"user_id": customer_id}},
        {"$group": {
            "_id": "$status",
            "count": {"$sum": 1}
        }}
    ]).to_list(10)
    
    return_summary = {
        "total_returns": sum(r["count"] for r in return_counts),
        "pending": next((r["count"] for r in return_counts if r["_id"] == "requested"), 0),
        "approved": next((r["count"] for r in return_counts if r["_id"] == "approved"), 0),
        "completed": next((r["count"] for r in return_counts if r["_id"] == "completed"), 0),
        "rejected": next((r["count"] for r in return_counts if r["_id"] == "rejected"), 0)
    }
    
    # Add is_active field if not present
    customer["is_active"] = customer.get("is_active", True)
    
    return {
        "profile": customer,
        "orders": orders,
        "returns": returns,
        "financial_summary": financial_summary,
        "return_summary": return_summary
    }


@router.put("/{customer_id}/status")
async def update_customer_status(customer_id: str, status_data: CustomerStatusUpdate, request: Request):
    """
    Activate or deactivate a customer account.
    Admin only.
    """
    await require_admin(request)
    
    # Check if customer exists and is not admin
    customer = await db.users.find_one({"user_id": customer_id, "role": {"$ne": "admin"}})
    
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    # Update status
    result = await db.users.update_one(
        {"user_id": customer_id},
        {"$set": {"is_active": status_data.is_active, "updated_at": datetime.now(timezone.utc)}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=400, detail="Failed to update customer status")
    
    return {
        "message": f"Customer {'activated' if status_data.is_active else 'deactivated'} successfully",
        "user_id": customer_id,
        "is_active": status_data.is_active
    }


@router.get("/export/csv")
async def export_customers_csv(
    request: Request,
    from_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    to_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    status: Optional[str] = Query(None, description="Filter by status: active, inactive, all")
):
    """
    Export customers to CSV file.
    Admin only.
    """
    await require_admin(request)
    
    # Build match filter
    match_filter = {
        "role": {"$ne": "admin"},
        "$or": [
            {"is_deleted": {"$exists": False}},
            {"is_deleted": False}
        ]
    }
    
    # Date filter
    if from_date:
        start = datetime.strptime(from_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
        match_filter["created_at"] = {"$gte": start}
    
    if to_date:
        end = datetime.strptime(to_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59, tzinfo=timezone.utc)
        if "created_at" in match_filter:
            match_filter["created_at"]["$lte"] = end
        else:
            match_filter["created_at"] = {"$lte": end}
    
    # Status filter
    if status == "active":
        match_filter["$and"] = match_filter.get("$and", []) + [{
            "$or": [
                {"is_active": {"$exists": False}},
                {"is_active": True}
            ]
        }]
    elif status == "inactive":
        match_filter["is_active"] = False
    
    # Aggregation pipeline
    pipeline = [
        {"$match": match_filter},
        {"$lookup": {
            "from": "orders",
            "localField": "user_id",
            "foreignField": "user_id",
            "as": "orders"
        }},
        {"$addFields": {
            "total_orders": {"$size": "$orders"},
            "total_spend": {
                "$sum": {
                    "$map": {
                        "input": {
                            "$filter": {
                                "input": "$orders",
                                "as": "order",
                                "cond": {"$ne": ["$$order.order_status", "cancelled"]}
                            }
                        },
                        "as": "o",
                        "in": "$$o.total"
                    }
                }
            },
            "last_order_date": {"$max": "$orders.created_at"},
            "is_active": {"$ifNull": ["$is_active", True]}
        }},
        {"$project": {
            "_id": 0,
            "user_id": 1,
            "name": 1,
            "email": 1,
            "phone": 1,
            "auth_provider": 1,
            "created_at": 1,
            "is_active": 1,
            "total_orders": 1,
            "total_spend": 1,
            "last_order_date": 1
        }},
        {"$sort": {"created_at": -1}}
    ]
    
    customers = await db.users.aggregate(pipeline).to_list(10000)
    
    # Create CSV
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Headers
    writer.writerow([
        "Customer ID",
        "Name",
        "Email",
        "Phone",
        "Auth Provider",
        "Status",
        "Total Orders",
        "Total Spend (₹)",
        "Last Order Date",
        "Registered On"
    ])
    
    # Data rows
    for customer in customers:
        last_order = customer.get("last_order_date")
        last_order_str = last_order.strftime("%Y-%m-%d %H:%M") if last_order else "Never"
        
        created_at = customer.get("created_at")
        created_str = created_at.strftime("%Y-%m-%d %H:%M") if created_at else ""
        
        writer.writerow([
            customer.get("user_id", ""),
            customer.get("name", ""),
            customer.get("email", ""),
            customer.get("phone", ""),
            customer.get("auth_provider", "email"),
            "Active" if customer.get("is_active", True) else "Inactive",
            customer.get("total_orders", 0),
            round(customer.get("total_spend", 0), 2),
            last_order_str,
            created_str
        ])
    
    output.seek(0)
    
    # Generate filename with date range
    filename = "customers_export"
    if from_date:
        filename += f"_from_{from_date}"
    if to_date:
        filename += f"_to_{to_date}"
    filename += ".csv"
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/excel")
async def export_customers_excel(
    request: Request,
    from_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    to_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    status: Optional[str] = Query(None, description="Filter by status: active, inactive, all")
):
    """
    Export customers to Excel file.
    Admin only.
    """
    await require_admin(request)
    
    # Try to import openpyxl, fall back to CSV if not available
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(
            status_code=501,
            detail="Excel export not available. Please use CSV export instead."
        )
    
    # Build match filter
    match_filter = {
        "role": {"$ne": "admin"},
        "$or": [
            {"is_deleted": {"$exists": False}},
            {"is_deleted": False}
        ]
    }
    
    # Date filter
    if from_date:
        start = datetime.strptime(from_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
        match_filter["created_at"] = {"$gte": start}
    
    if to_date:
        end = datetime.strptime(to_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59, tzinfo=timezone.utc)
        if "created_at" in match_filter:
            match_filter["created_at"]["$lte"] = end
        else:
            match_filter["created_at"] = {"$lte": end}
    
    # Status filter
    if status == "active":
        match_filter["$and"] = match_filter.get("$and", []) + [{
            "$or": [
                {"is_active": {"$exists": False}},
                {"is_active": True}
            ]
        }]
    elif status == "inactive":
        match_filter["is_active"] = False
    
    # Aggregation pipeline
    pipeline = [
        {"$match": match_filter},
        {"$lookup": {
            "from": "orders",
            "localField": "user_id",
            "foreignField": "user_id",
            "as": "orders"
        }},
        {"$addFields": {
            "total_orders": {"$size": "$orders"},
            "total_spend": {
                "$sum": {
                    "$map": {
                        "input": {
                            "$filter": {
                                "input": "$orders",
                                "as": "order",
                                "cond": {"$ne": ["$$order.order_status", "cancelled"]}
                            }
                        },
                        "as": "o",
                        "in": "$$o.total"
                    }
                }
            },
            "last_order_date": {"$max": "$orders.created_at"},
            "is_active": {"$ifNull": ["$is_active", True]}
        }},
        {"$project": {
            "_id": 0,
            "user_id": 1,
            "name": 1,
            "email": 1,
            "phone": 1,
            "auth_provider": 1,
            "created_at": 1,
            "is_active": 1,
            "total_orders": 1,
            "total_spend": 1,
            "last_order_date": 1
        }},
        {"$sort": {"created_at": -1}}
    ]
    
    customers = await db.users.aggregate(pipeline).to_list(10000)
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="E10600", end_color="E10600", fill_type="solid")
    
    # Headers
    headers = [
        "Customer ID",
        "Name",
        "Email",
        "Phone",
        "Auth Provider",
        "Status",
        "Total Orders",
        "Total Spend (₹)",
        "Last Order Date",
        "Registered On"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    for row, customer in enumerate(customers, 2):
        last_order = customer.get("last_order_date")
        last_order_str = last_order.strftime("%Y-%m-%d %H:%M") if last_order else "Never"
        
        created_at = customer.get("created_at")
        created_str = created_at.strftime("%Y-%m-%d %H:%M") if created_at else ""
        
        ws.cell(row=row, column=1, value=customer.get("user_id", ""))
        ws.cell(row=row, column=2, value=customer.get("name", ""))
        ws.cell(row=row, column=3, value=customer.get("email", ""))
        ws.cell(row=row, column=4, value=customer.get("phone", ""))
        ws.cell(row=row, column=5, value=customer.get("auth_provider", "email"))
        ws.cell(row=row, column=6, value="Active" if customer.get("is_active", True) else "Inactive")
        ws.cell(row=row, column=7, value=customer.get("total_orders", 0))
        ws.cell(row=row, column=8, value=round(customer.get("total_spend", 0), 2))
        ws.cell(row=row, column=9, value=last_order_str)
        ws.cell(row=row, column=10, value=created_str)
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename
    filename = "customers_export"
    if from_date:
        filename += f"_from_{from_date}"
    if to_date:
        filename += f"_to_{to_date}"
    filename += ".xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
